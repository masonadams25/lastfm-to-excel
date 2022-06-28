import openpyxl

'''
c_1 = sheet.cell(row = 1, column = 1)
c_2 = sheet.cell(row = 2, column = 1)
c_3 = sheet.cell(row = 2, column = 1)

c_1.value = "TEST1"
c_2.value = "TEST2"
c_3.value = "TEST3"

wb.save(path)

print("One: " + c_1.value)
print("Two: " + c_2.value)
print("Three: " + c_3.value)
'''

def print_to_excel(x, artist, count):
    path = "py_test.xlsx"

    wb = openpyxl.load_workbook(path)

    sheet = wb.active

    (sheet.cell(row = x+1, column = 1)).value = f"{artist}"
    (sheet.cell(row = x+1, column = 2)).value = f"{count}"

    wb.save(path)

