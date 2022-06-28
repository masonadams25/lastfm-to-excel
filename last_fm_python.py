import openpyxl

path = "/Users/masonadams/Downloads/py_test.xlsx"

wb = openpyxl.load_workbook(path)

sheet = wb.active


'''
c_1 = sheet.cell(row = 1, column = 1)
c_2 = sheet.cell(row = 2, column = 1)
c_3 = sheet.cell(row = 2, column = 1)

c_1.value = "TEST1"
c_2.value = "TEST2"
c_3.value = "TEST3"

wb.save(path)

print("One: " + c_1.value)
print("Two: " + c_2.value)
print("Three: " + c_3.value)
'''

for x in range(100):
    for y in range(100):
        (sheet.cell(row = x+1, column = y+1)).value = f"Cell: {x}-{y}"

wb.save(path)
